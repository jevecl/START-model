import os
import gurobipy as gp
from gurobipy import GRB
import pandas as pd
import numpy as np
import openpyxl
from STARTmodel import START

from utils import create_color_palettes, apply_color, access_model_variables, add_sheet_excel, columns_dimensions, sorted_profiles
#  Read the data from the excel file
#datafile='Data.xlsx'
#datafile='Data2Inf.xlsx'
datafile='Simulate_data_small.xlsx'
#datafile='Simulate_data_small.xlsx'

shdata = pd.read_excel(datafile, sheet_name='Data', header=None)
pNperiods = shdata.loc[0,1]
pNhealthp = shdata.loc[1,1]
pNpeople = shdata.loc[2,1]
pNcharter = shdata.loc[3,1]
pDiscount = shdata.loc[4,1]
pKpeople = shdata.loc[5,1]
pMaxperiods = shdata.loc[6,1]
pMinperiods = shdata.loc[7,1]

shprices = pd.read_excel(datafile, sheet_name='Prices')
pCostOut = shprices.loc[:,'Outward'].to_numpy()
pCostRet = shprices.loc[:,'Return'].to_numpy()

shcharter = pd.read_excel(datafile, sheet_name='Charter')
pCostChar = shcharter.loc[0:pNperiods-1,'Chartered 1':'Chartered 2'].to_numpy()
pMinCapCh = shcharter.loc[0:pNcharter-1,'Min Cap'].to_numpy()
pMaxCapCh = shcharter.loc[0:pNcharter-1,'Max Cap'].to_numpy()

shhealthprofiles = pd.read_excel(datafile, sheet_name='HealthProfiles', header=None)
pNameprofiles = shhealthprofiles.loc[0,1:].to_numpy()
pNameabbprofiles = shhealthprofiles.loc[1,1:].to_numpy()
pProfiles = shhealthprofiles.loc[2:,1:].to_numpy()

def create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg):
    """
    Create the model for the problem.
    """
    model = START(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)
    model.create_model()
    model.create_variables()
    model.create_constrains()
    model.create_objective_function()
    solution = model.solve()
    return model, solution

def model_dimensions(model, excel):
    """
    Create the sheet 'Model dimensions' in the excel file.
    The sheet contains the number of variables, constraints and the type of variables.
    """
    df = pd.DataFrame(index = ['total', 'continuous', 'binary', 'integer', 'constrains'], columns=['Type', 'Number'])
    df['Type'] = df.index
    df['Number'] = [    
        model.NumVars,
        model.NumVars-model.NumIntVars,
        model.NumBinVars,
        model.NumIntVars-model.NumBinVars, 
        model.NumConstrs
    ]
    add_sheet_excel(excel, 'Model dimensions', df, index = True)
    wb = openpyxl.load_workbook(excel)
    sheet = wb['Model dimensions']
    columns_dimensions(excel, wb, sheet, df, 25)

def attributes_values(model, excel_file): 
    """
    Create the sheet 'Attributes values' in the excel file. The sheet contains the values of the attributes of the model.
    """
    df = pd.DataFrame(index = ['Number of people', 'Cost' ,'Availability', 'Grade'], columns=['Value'])
    meanav = round(model.vMeanav.X, 2)
    meangrade = round(model.vMeangrade.X, 2)
    df['Value'] = [model.vTotpeop.X,
        model.vCost.X,
        meanav,
        meangrade
    ]
    add_sheet_excel(excel_file, 'Attributes values', df, index = True)
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Attributes values']
    columns_dimensions(excel_file, wb, sheet, df, 25)

def model_performance(model,excel): 
    info = ['Objective value', 'Runtime', 'MIPGap', 'Status']
    df = pd.DataFrame(index = info, columns=['Value'])
    df['Value'] = [
        model.objVal,
        model.Runtime,
        model.MIPGap,
        model.Status
    ]
    add_sheet_excel(excel, 'Model performance', df, index = True)
    wb = openpyxl.load_workbook(excel)
    sheet = wb['Model performance']
    columns_dimensions(excel, wb, sheet, df, 25)

def chartered(model, excel):
    """
    Create the sheet 'Charter' in the excel file. The sheet contains the number of people that are chartered and the cost.
    """ 
    vGamma = access_model_variables('vGamma', 2, model)
    charter_data = []
    cost_data = []
    vGamma_array = np.array([[vGamma[t, l].X for l in range(pNcharter)] for t in range(pNperiods)])
    for l in range(pNcharter):
        nl = np.sum(vGamma_array[:, l])
        costl = np.sum(pCostChar[:, l] * vGamma_array[:, l])
        charter_data.append(nl)
        cost_data.append(costl)
    df = pd.DataFrame(index = ['Chartered 1', 'Chartered 2'], columns=['Number', 'Cost'])
    df['Number'] = charter_data
    df['Cost'] = cost_data
    add_sheet_excel(excel, 'Charter', df, index = True)

def regular_and_cost(excel, model): 
    """
    Create the sheet 'Regular' in the excel file. The sheet contains the total cost of the outward and return flights and the 
    sum of both.
    """
    costoutflights = 0
    costretflights = 0
    vXstand = access_model_variables('Xstand', 1, model)
    vXdisc = access_model_variables('Xdisc', 1, model)
    vYstand = access_model_variables('Ystand', 1, model)
    vYdisc = access_model_variables('Ydisc', 1, model)
    for t in range(pNperiods - 1):
        costoutflights += pCostOut[t] * vXstand[t].X + (pCostOut[t] - pDiscount) * vXdisc[t].X
        costretflights += pCostRet[t + 1] * vYstand[t + 1].X + (pCostRet[t + 1] - pDiscount) * vYdisc[t + 1].X
    totcostreg = costoutflights + costretflights
    df = pd.DataFrame(index = ['Outward', 'Return','Total'], columns=['Value'])
    df['Value'] = [costoutflights, costretflights, totcostreg]
    add_sheet_excel(excel, 'Regular', df, index = True)

def out_and_return(excel, model):
    """
    Create the sheet 'Outwards and return' in the excel file. The sheet contains the number of people that are
    outwards and return depending on the type of flight.
    """
    vXstand = access_model_variables('Xstand', 1, model)
    vXdisc = access_model_variables('Xdisc', 1, model)
    vYstand = access_model_variables('Ystand', 1, model)
    vYdisc = access_model_variables('Ydisc', 1, model)
    vZout = access_model_variables('Zout', 1, model)
    vZret = access_model_variables('Zret', 1, model) 

    npeopstand = sum(vXstand[t].X for t in range(pNperiods-1))
    npeopdisc = sum(vXdisc[t].X for t in range(pNperiods-1))
    npeopchar = sum(vZout[t].X for t in range(pNperiods-1))

    npeopstandr = sum(vYstand[t].X for t in range(1, pNperiods))
    npeopdiscr = sum(vYdisc[t].X for t in range(1, pNperiods))
    npeopcharr = sum(vZret[t].X for t in range(1, pNperiods))

    df = pd.DataFrame(index = ['Standard', 'Discounted', 'Chartered'], columns=['Number of people outwards', 'Number of people return'])
    df['Number of people outwards'] = [npeopstand, npeopdisc, npeopchar]
    df['Number of people return'] = [npeopstandr, npeopdiscr, npeopcharr]
    add_sheet_excel(excel, 'Outwards and return', df, index = True)
 
def fligths_plan(excel, model): 
    """
    Create the sheet 'Flights plan' in the excel file. The sheet contains the period of departure and return of each person.
    """
    vAlphaout = access_model_variables('Alphaout', 2, model)
    vAlpharet = access_model_variables('Alpharet', 2, model)
    dict = {}
    perdep = 0
    perret = 0
    for i in range(0, pNpeople):
        for t in range(0, pNperiods):
            if (vAlphaout[i,t].X == 1):
                perdep = t
                for tr in range(t+1, pNperiods):
                    if (vAlpharet[i,tr].X == 1):
                            perret = tr
                            dict[f'Person {i+1:03d}'] = [perdep+1, perret]
    df_people = pd.DataFrame.from_dict(dict, orient='index', columns=['Departure', 'Return'])
    df_people['Person'] =  dict.keys()
    df_people = df_people[['Person', 'Departure', 'Return']]  
    add_sheet_excel(excel, 'Flights plan', df_people, index = False)
    wb = openpyxl.load_workbook(excel)
    sheet = wb['Flights plan']
    columns_dimensions(excel, wb, sheet, df_people, 25)

def health_profile_complementary(excel_file ,df, dict_references, color_palettes):
    """
    Create the sheet 'Profile plan' in the excel file. The sheet contains the health profile of each person in each period.
    """
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.create_sheet('Profile plan')
    for i in range(1, df.shape[1]):
        sheet.cell(row=1, column=i + 1).value = 'Period ' + str(i)
    for i, index_value in enumerate(df.index, start=1):
         sheet.cell(row=i + 1, column=1).value = 'Person ' + str(index_value).zfill(3)
    for i, row in enumerate(df.values):
        for j, cell_value in enumerate(row):
            color_fill = apply_color(cell_value, color_palettes)
            if cell_value == 0:
                sheet.cell(row=i + 2, column=j + 2).value = ''
            else:
                sheet.cell(row=i + 2, column=j + 2).value = dict_references[cell_value] 
                sheet.cell(row=i + 2, column=j + 2).fill = color_fill
    for i, key in enumerate(dict_references):
        sheet.cell(row=i + 5, column=df.shape[1] + 3).value = dict_references[key]
        color_fill = apply_color(key, color_palettes)
        sheet.cell(row=i + 5, column=df.shape[1] + 2).fill = color_fill
    workbook.save(excel_file)

def health_profiles(model, excel_file):
    """
    Create the sheet 'Health profiles' in the excel file. The sheet contains the health profile of each person in each period.
    """
    results = np.zeros((pNpeople, pNperiods), dtype=int)
    vBeta = access_model_variables('vBeta', 3, model)
    for i in range(0,pNpeople):
        for t in range(0,pNperiods):
            results[i,t] = 0
            for j in range (0,pNhealthp):
                if vBeta[i,j,t].X == 1:
                    results[i,t] = j+1
    dict_health_profiles = dict(zip(range(1, pNhealthp + 1), pNameabbprofiles))
    df = pd.DataFrame(results, index=range(1, pNpeople + 1))  
    df = df.loc[~(df==0).all(axis=1)]
    color_palettes = create_color_palettes(pNhealthp)
    df_sorted = sorted_profiles(df, pNperiods)
    health_profile_complementary(excel_file, df_sorted, dict_health_profiles, color_palettes)
    
def necessary_people(excel_file, model): 
    """
    Create the sheet 'Necessary people' in the excel file. The sheet contains the number of complementary people that are necessary to 
    attend the emergency in each period.
    """
    vUmas = access_model_variables('vUmas', 2, model)
    df = pd.DataFrame(index = range(0, pNhealthp), columns=range(0, pNperiods))
    dict_health_profiles = dict(zip(range(1, pNhealthp + 1), pNameabbprofiles))
    df_values = np.array([[vUmas[j, t].X for t in range(pNperiods)] for j in range(pNhealthp)])
    df = pd.DataFrame(df_values)
    df.index = dict_health_profiles.values()
    for i in range(0, pNperiods):
        df.rename(columns={i: 'Period ' + str(i+1)}, inplace=True)
    add_sheet_excel(excel_file, 'Necessary people', df, index = True)

def payoff_matrix(excel_file, array_payoff):
    """
    Show the payoff matrix of the problem
    """
    df = pd.DataFrame(array_payoff)
    df.iloc[:, 1] = df.iloc[:, 1].round(2)
    df.iloc[:, 2] = df.iloc[:, 2].round(2)
    df.index = ['Cost', 'Availability', 'Grades'] 
    df.columns = ['Cost', 'Availability', 'Grades']
    add_sheet_excel(excel_file, 'Payoff Matrix', df, index = True)

def weights_goals(excel_file, array_weights, arrays_goals = None):
    """
    Show the weights of the goal programming and compromise programming
    """
    df = pd.DataFrame(array_weights)
    df.index = ['Cost', 'Availability', 'Grades']
    df.columns = ['Weights']
    if arrays_goals is not None:
        df['Goals'] = arrays_goals
        add_sheet_excel(excel_file, 'Weights and goals', df, index = True)
    else:
        add_sheet_excel(excel_file, 'Weights', df, index = True)

def create_payoff_matrix(model, payoff, type_m):
    """
    Create the payoff matrix of the problem
    """
    values = [model.vCost.X, model.vMeanav.X, model.vMeangrade.X]
    type_to_index = {"cost": 0, "availability": 1, "grades": 2}
    
    if type_m in type_to_index:
        index = type_to_index[type_m]
        payoff[index, :] = values

def print_model_performance(model, type_m):
    pInfeas = model.model.objVal
    npeople = model.vTotpeop.X
    cost = model.vCost.X
    avai = model.vMeanav.X
    grad = model.vMeangrade.X
    objfun = model.model.objVal
    comtim = model.model.Runtime
    gapmod = model.model.MIPGap
    status = model.model.Status
    print("==========================")
    if type_m == "infeasibility":
        print("INFEASIBILITY OPTIMIZATION")
    elif type_m == "cost":
        print("COST OPTIMIZATION")
    elif type_m == "availability":
        print("AVAILABILITY OPTIMIZATION")
    elif type_m == "grades":
        print("GRADES OPTIMIZATION")
    elif type_m == "goal":
        print("GOAL PROGRAMMING")
    elif type_m == "monobjective":
        print("MONOOBJECTIVE PROGRAMMING")
    elif type_m == "compromise":
        print("COMPROMISE PROGRAMMING")
    print("==========================")
    
    if type_m == "infeasibility":
        print(f"Objective function: {objfun}")
    else: 
        print(f"Number of people: {npeople}")
        print(f"Cost:             {cost}")
        print(f"Availability:     {avai}")
        print(f"Grades:           {grad}")
        print(f"Objective function: {objfun}")
        print(f"Computing time    : {comtim}")
        print(f"Optimality gap    : {gapmod}")
        print(f"Status of solution: {status}")

def weights(excel_file, model_t):
    """
    Read the weights for resolving a type of problem and put them in a dataframe.
    The weights are read from a column specified by 'model_t' in the Excel file.
    """
    # Initialize an empty DataFrame with specified indices
    df = pd.DataFrame(index=['pWC', 'pWA', 'pWG', 'pWO', 'pWGC', 'pWGA', 'pWGG', 'pGoalc', 'pGoala', 'pGoalg'],
                      columns=[model_t])
    
    # Read the specific column from the Excel file, assuming the column name matches 'model_t'
    # The function 'usecols' parameter can specify a single column to read
    # 'nrows' parameter is set according to the expected number of weights
    weight_data = pd.read_excel(excel_file, usecols=[model_t], nrows=len(df.index))
    
    # Assign the read weights to the DataFrame under the column named after 'model_t'
    df[model_t] = weight_data.values.flatten()  # Flatten the array if necessary

    return df
 

def main(filename_res, model, solution, list_goals = None, list_weights = None):    
    excel_file_path = os.path.join(os.getcwd(), filename_res)
    wb = openpyxl.Workbook()
    wb.active.title = 'Model dimensions'
    wb.save(excel_file_path)
    model_performance(solution, excel_file_path)
    model_dimensions(solution, excel_file_path)
    attributes_values(model, excel_file_path)
    chartered(solution, excel_file_path)
    regular_and_cost(excel_file_path, solution)
    out_and_return(excel_file_path, solution)
    fligths_plan(excel_file_path, solution)
    health_profiles(solution, excel_file_path)   
    necessary_people(excel_file_path, solution)
    if "GP1" in filename_res or "GP2" in filename_res or "Compromise" in filename_res or "Monobjective" in filename_res:
        payoff_matrix(excel_file_path, payoff)
        weights_goals(excel_file_path, list_weights, list_goals)


if __name__ == '__main__':
    os.makedirs('results', exist_ok=True)

    vBetaini = {}
    pMaxgrade = 10

    pIdc = 0
    pIda = 0
    pIdg = 0

    # First resolution is for testing feasibility, then, pInfeas is high
    pInfeas = 9999
    pWI = 1
    pWC = 0
    pWA = 0
    pWG = 0
    pWO = 0
    pWGC = 0
    pWGA = 0
    pWGG = 0
    pGoalc = 0
    pGoala = 0
    pGoalg = 0
    df = weights('weights.xlsx', 'infeseability')
    print(df)
    
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)
    print_model_performance(model, "infeasibility")
    filename_infeas = os.path.join('results', '01_Res_Simulate_data_Infeasibility.xlsx')
    main(filename_res=filename_infeas, model=model, solution=solution)

    # PAYOFF MATRIX
    payoff = np.empty((3, 3))

    # Minimizing only cost
    pWI = 0
    pWC = 1
    pWA = 0
    pWG = 0
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)
    for i in range(0, pNpeople):
        for j in range(0, pNhealthp):
            for t in range(0, pNperiods):
                vBetaini[i,j,t] = model.vBeta[i,j,t].X
    
    create_payoff_matrix(model, payoff, "cost")
    print_model_performance(model, "cost")
    filename_cost = os.path.join('results', '02_Res_Simulate_data_Cost.xlsx')
    main(filename_res=filename_cost, model=model, solution=solution)

    # Minimizing availability
    pWC = 0
    # As we want to maximize the availability, the weight is -1:
    pWA = 1
    pWG = 0

    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)
    create_payoff_matrix(model, payoff, "availability")
    print_model_performance(model, "availability")
    filename_availability = os.path.join('results', '03_Res_Simulate_data_Availability.xlsx')
    main(filename_res=filename_availability, model=model, solution=solution)

    # Minimizing grades
    pWC = 0
    pWA = 0
    # As we want to maximize the grades, the weight is -1:
    pWG = 1
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)
    create_payoff_matrix(model, payoff, "grades")
    print_model_performance(model, "grades")
    filename_grades = os.path.join('results', '04_Res_Simulate_data_Grades.xlsx')
    main(filename_res=filename_grades, model=model, solution=solution)

    # MONOOBJECTIVE RESOLUTION
    pWC = 1 / (max(payoff[1, 0], payoff[2, 0]) - payoff[0, 0])
    pWA = 1 / (payoff[1, 1] - min(payoff[0, 1], payoff[2, 1]))
    pWG = 1 / (payoff[2, 2] - min(payoff[0, 2], payoff[1, 2]))
    print(f"pWC:{pWC} pWA: {pWA}, pWG: {pWG}")
    # Defining weights for goals
    pWGC = 0
    pWGA = 0
    pWGG = 0

    # Defining goals
    pGoalc = 0
    pGoala = 0
    pGoalg = 0
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg,
                                   vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)

    print_model_performance(model, "monobjective")
    filename_monobjective = os.path.join('results', '05_Res_Simulate_data_Monoobjective.xlsx')
    main(filename_res=filename_monobjective, model=model, solution=solution, list_weights=[pWC, pWA, pWG])

    # GOAL PROGRAMMING RESOLUTION 1
    pPercost = 0.10
    pPerav = 0.10
    pPergr = 0.10

    pWC = 0
    pWA = 0
    pWG = 0
    pWO = 0

    # Defining weights for goals
    pWGC = 1/(max(payoff[1,0],payoff[2,0])-payoff[0,0])
    pWGA = 1/(payoff[1,1]-min(payoff[0,1],payoff[2,1]))
    pWGG = 1/(payoff[2,2]-min(payoff[0,2],payoff[1,2]))
    print(f"pWGC:{pWGC} pWGA: {pWGA}, pWGG: {pWGG}")
    # Defining goals
    pGoalc = payoff[0,0] * (1+pPercost)
    #pGoala = payoff[1,1]*(1-pPercost)
    pGoala = payoff[1,1] - pPerav * (payoff[1,1] - min(payoff[0,1], payoff[2,1]))
    #pGoalg = payoff[2,2]*(1-pPercost)
    pGoalg = payoff[2,2] - pPergr * (payoff[2,2] - min(payoff[0,2], payoff[1,2]))
    print(f"GoalC:{pGoalc} GoalA: {pGoala}, GoalG: {pGoalg}")
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)

    npeople = model.vTotpeop.X
    cost = model.vCost.X
    avai = model.vMeanav.X
    grad = model.vMeangrade.X

    print("==================")
    print("GOAL PROGRAMMING 1")
    print("==================")
    print(f"Number of people: {npeople}")
    print(f"Cost:             {cost}")
    print(f"Availability:     {avai}")
    print(f"Grades:           {grad}")
    objfun = model.model.objVal
    comtim = model.model.Runtime
    gapmod = model.model.MIPGap
    status = model.model.Status

    print(f"Objective function: {objfun}")
    print(f"Computing time    : {comtim}")
    print(f"Optimality gap    : {gapmod}")
    print(f"Status of solution: {status}")

    filename_goal1 = os.path.join('results', '06_Res_Simulate_data_GP1.xlsx')
    main(filename_res=filename_goal1, model = model, solution = solution, list_goals = [pGoalc, pGoala, pGoalg], list_weights = [pWGC, pWGA, pWGG])

    # GOAL PROGRAMMING RESOLUTION 2
    pPercost = 0.05
    pPerav = 0.05
    pPergr = 0.05

    pWC = 0
    pWA = 0
    pWG = 0
    pWO = 0

    # Defining weights for goals
    pWGC = 1 / (max(payoff[1, 0], payoff[2, 0]) - payoff[0, 0])
    pWGA = 1 / (payoff[1, 1] - min(payoff[0, 1], payoff[2, 1]))
    pWGG = 1 / (payoff[2, 2] - min(payoff[0, 2], payoff[1, 2]))
    print(f"pWGC:{pWGC} pWGA: {pWGA}, pWGG: {pWGG}")
    # Defining goals
    pGoalc = payoff[0, 0] * (1 + pPercost)
    # pGoala = payoff[1,1]*(1-pPercost)
    pGoala = payoff[1, 1] - pPerav * (payoff[1, 1] - min(payoff[0, 1], payoff[2, 1]))
    # pGoalg = payoff[2,2]*(1-pPercost)
    pGoalg = payoff[2, 2] - pPergr * (payoff[2, 2] - min(payoff[0, 2], payoff[1, 2]))
    print(f"GoalC:{pGoalc} GoalA: {pGoala}, GoalG: {pGoalg}")
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg,
                                   vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)

    npeople = model.vTotpeop.X
    cost = model.vCost.X
    avai = model.vMeanav.X
    grad = model.vMeangrade.X

    print("==================")
    print("GOAL PROGRAMMING 2")
    print("==================")
    print(f"Number of people: {npeople}")
    print(f"Cost:             {cost}")
    print(f"Availability:     {avai}")
    print(f"Grades:           {grad}")
    objfun = model.model.objVal
    comtim = model.model.Runtime
    gapmod = model.model.MIPGap
    status = model.model.Status

    print(f"Objective function: {objfun}")
    print(f"Computing time    : {comtim}")
    print(f"Optimality gap    : {gapmod}")
    print(f"Status of solution: {status}")

    filename_goal2 = os.path.join('results', '07_Res_Simulate_data_GP2.xlsx')
    main(filename_res=filename_goal2, model=model, solution=solution, list_goals=[pGoalc, pGoala, pGoalg],
         list_weights=[pWGC, pWGA, pWGG])

    # COMPROMISE PROGRAMMING RESOLUTION
    pWC = 1 / (max(payoff[1, 0], payoff[2, 0]) - payoff[0, 0])
    pWA = 1 / (payoff[1, 1] - min(payoff[0, 1], payoff[2, 1]))
    pWG = 1 / (payoff[2, 2] - min(payoff[0, 2], payoff[1, 2]))
    print(f"pWC:{pWC} pWA: {pWA}, pWG: {pWG}")
    # Defining weights for goals
    pWGC = 0
    pWGA = 0
    pWGG = 0

    pIdc = payoff[0,0]
    pIda = payoff[1,1]
    pIdg = payoff[2,2]

    # Defining goals
    pGoalc = 0
    pGoala = 0
    pGoalg = 0
    model, solution = create_model(datafile, pWI, pWC, pWA, pWO, pWG, pInfeas, pWGC, pWGA, pWGG, pGoalc, pGoala, pGoalg, vBetaini, pNpeople, pNperiods, pNhealthp, pMaxgrade, pIdc, pIda, pIdg)

    npeople = model.vTotpeop.X
    cost = model.vCost.X
    avai = model.vMeanav.X
    grad = model.vMeangrade.X

    print("======================")
    print("COMPROMISE PROGRAMMING")
    print("======================")
    print(f"Number of people: {npeople}")
    print(f"Cost:             {cost}")
    print(f"Availability:     {avai}")
    print(f"Grades:           {grad}")

    objfun = model.model.objVal
    comtim = model.model.Runtime
    gapmod = model.model.MIPGap
    status = model.model.Status

    print(f"Objective function: {objfun}")
    print(f"Computing time    : {comtim}")
    print(f"Optimality gap    : {gapmod}")
    print(f"Status of solution: {status}")

    filename_compromise = os.path.join('results', '08_Res_Simulate_data_Compromise.xlsx')
    main(filename_res= filename_compromise, model=model, solution = solution, list_weights = [pWC, pWA, pWG])